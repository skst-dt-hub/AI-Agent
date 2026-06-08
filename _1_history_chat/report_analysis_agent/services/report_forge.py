from __future__ import annotations

import json
import re
from dataclasses import asdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from config import DATA_SOURCE_ID, KNOWLEDGE_BASE_ID, MODEL_ID, OUTPUT_DIR
from models import AnalysisRun


def save_outputs(run: AnalysisRun) -> AnalysisRun:
    run_dir = _make_run_dir(run.keyword, run.started_at)
    run_dir.mkdir(parents=True, exist_ok=True)

    markdown_path = run_dir / "report.md"
    markdown_path.write_text(run.markdown_report, encoding="utf-8")

    excel_path = run_dir / "report.xlsx"
    run.output_dir = run_dir
    run.markdown_path = markdown_path
    run.excel_path = excel_path
    _write_excel(run, excel_path)

    metadata_path = run_dir / "run_metadata.json"
    metadata_path.write_text(
        json.dumps(_metadata(run, markdown_path, excel_path), ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )

    return run


def _make_run_dir(keyword: str, started_at: datetime) -> Path:
    safe_keyword = _safe_filename(keyword) or "analysis"
    timestamp = started_at.strftime("%Y%m%d_%H%M%S")
    return Path(OUTPUT_DIR) / safe_keyword / timestamp


def _safe_filename(value: str) -> str:
    value = re.sub(r"[^\w.-]+", "_", value.strip(), flags=re.UNICODE)
    return value.strip("._")[:80]


def _write_excel(run: AnalysisRun, path: Path) -> None:
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    _write_summary(summary, run)

    internal = wb.create_sheet("Internal")
    _write_internal(internal, run)

    external = wb.create_sheet("External")
    _write_external(external, run)

    report = wb.create_sheet("Report")
    _write_report(report, run)

    log = wb.create_sheet("RunLog")
    _write_runlog(log, run)

    for ws in wb.worksheets:
        _style_sheet(ws)

    wb.save(path)


def _write_summary(ws, run: AnalysisRun) -> None:
    rows = [
        ("Keyword", run.keyword),
        ("Started At", run.started_at.isoformat(timespec="seconds")),
        ("Scout Status", _step_text(run.scout)),
        ("Ranger Status", _step_text(run.ranger)),
        ("Anchor Status", _step_text(run.anchor)),
        ("Internal Results", len(run.internal_results)),
        ("External Results", len(run.external_results)),
    ]
    for row in rows:
        ws.append(row)


def _write_internal(ws, run: AnalysisRun) -> None:
    ws.append(["No", "Score", "Source", "Content", "Metadata"])
    for index, item in enumerate(run.internal_results, start=1):
        ws.append(
            [
                index,
                item.score,
                item.source,
                item.content,
                json.dumps(item.metadata, ensure_ascii=False, default=str),
            ]
        )


def _write_external(ws, run: AnalysisRun) -> None:
    ws.append(["No", "Date", "Source", "Title", "URL", "Summary"])
    for index, item in enumerate(run.external_results, start=1):
        ws.append([index, item.date, item.source, item.title, item.url, item.summary])


def _write_report(ws, run: AnalysisRun) -> None:
    ws.append(["Line", "Content"])
    for index, line in enumerate(run.markdown_report.splitlines(), start=1):
        ws.append([index, line])


def _write_runlog(ws, run: AnalysisRun) -> None:
    rows: list[tuple[str, Any]] = [
        ("Model ID", MODEL_ID),
        ("Knowledge Base ID", KNOWLEDGE_BASE_ID),
        ("Data Source ID", DATA_SOURCE_ID),
        ("Output Dir", str(run.output_dir or "")),
        ("Markdown Path", str(run.markdown_path or "")),
        ("Excel Path", str(run.excel_path or "")),
        ("Scout", asdict(run.scout)),
        ("Ranger", asdict(run.ranger)),
        ("Anchor", asdict(run.anchor)),
    ]
    for key, value in rows:
        if isinstance(value, dict):
            value = json.dumps(value, ensure_ascii=False)
        ws.append([key, value])


def _style_sheet(ws) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="E8EEF7")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for column_cells in ws.columns:
        column_letter = column_cells[0].column_letter
        max_length = max(len(str(cell.value or "")) for cell in column_cells[:100])
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 80)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _step_text(step) -> str:
    detail = step.message or step.error
    return f"{step.status}: {detail}" if detail else step.status


def _metadata(run: AnalysisRun, markdown_path: Path, excel_path: Path) -> dict[str, Any]:
    return {
        "keyword": run.keyword,
        "started_at": run.started_at.isoformat(timespec="seconds"),
        "model_id": MODEL_ID,
        "knowledge_base_id": KNOWLEDGE_BASE_ID,
        "data_source_id": DATA_SOURCE_ID,
        "scout": asdict(run.scout),
        "ranger": asdict(run.ranger),
        "anchor": asdict(run.anchor),
        "search_plan": asdict(run.search_plan) if run.search_plan else None,
        "internal_results": [
            {"no": index, **asdict(item)}
            for index, item in enumerate(run.internal_results, start=1)
        ],
        "external_results": [
            {"no": index, **asdict(item)}
            for index, item in enumerate(run.external_results, start=1)
        ],
        "search_logs": [asdict(item) for item in run.search_logs],
        "markdown_path": str(markdown_path),
        "excel_path": str(excel_path),
    }

