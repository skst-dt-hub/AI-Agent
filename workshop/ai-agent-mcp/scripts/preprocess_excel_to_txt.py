from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Convert Excel workbooks into deterministic TXT files for KB ingestion."
    )
    parser.add_argument("--input", required=True, help="Input xlsx file or directory.")
    parser.add_argument("--output", required=True, help="Output directory for generated TXT files.")
    parser.add_argument("--include-hidden", action="store_true", help="Include hidden and veryHidden sheets.")
    parser.add_argument("--ascii-names", action="store_true", help="Use ASCII output paths while preserving original names in TXT content.")
    parser.add_argument("--max-file-kb", type=int, default=100, help="Keep sheet-level TXT if <= this size; split larger sheets by rows.")
    parser.add_argument("--split-overlap-rows", type=int, default=3, help="Rows to overlap between split parts for large sheets.")
    parser.add_argument("--verify-term", action="append", default=[], help="Term to count in generated TXT. Can be specified multiple times.")
    args = parser.parse_args()

    input_path = Path(args.input).resolve()
    output_dir = Path(args.output).resolve()
    files = collect_xlsx_files(input_path)
    if not files:
        raise SystemExit(f"No .xlsx files found: {input_path}")

    output_dir.mkdir(parents=True, exist_ok=True)
    manifest: list[dict[str, Any]] = []
    for file_index, file_path in enumerate(files, start=1):
        manifest.extend(
            convert_workbook(
                file_index=file_index,
                file_path=file_path,
                output_dir=output_dir,
                include_hidden=args.include_hidden,
                ascii_names=args.ascii_names,
                max_file_bytes=args.max_file_kb * 1024,
                split_overlap_rows=max(0, args.split_overlap_rows),
            )
        )

    manifest_path = output_dir / "manifest.json"
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"workbooks={len(files)}")
    print(f"txt_files={len(manifest)}")
    print(f"manifest={manifest_path}")
    for term in args.verify_term:
        print(f"verify_term[{term}]={count_term(output_dir, term)}")
    return 0


def collect_xlsx_files(input_path: Path) -> list[Path]:
    if input_path.is_file() and input_path.suffix.lower() == ".xlsx":
        return [input_path]
    if input_path.is_dir():
        return sorted(path for path in input_path.rglob("*.xlsx") if not path.name.startswith("~$"))
    return []


def convert_workbook(
    file_index: int,
    file_path: Path,
    output_dir: Path,
    include_hidden: bool,
    ascii_names: bool,
    max_file_bytes: int,
    split_overlap_rows: int,
) -> list[dict[str, Any]]:
    workbook = load_workbook(file_path, data_only=True, read_only=False)
    department = infer_department(file_path)
    report_date = infer_report_date(file_path.name)
    workbook_dir = build_workbook_dir(output_dir, file_index, file_path, department, report_date, ascii_names)
    workbook_dir.mkdir(parents=True, exist_ok=True)

    entries: list[dict[str, Any]] = []
    try:
        for sheet_index, sheet in enumerate(workbook.worksheets, start=1):
            sheet_state = getattr(sheet, "sheet_state", "visible")
            if sheet_state != "visible" and not include_hidden:
                continue

            rows = extract_rows(sheet)
            if not rows:
                continue

            full_text = render_text(
                source_file=file_path,
                department=department,
                report_date=report_date,
                sheet_name=sheet.title,
                sheet_state=sheet_state,
                row_start=rows[0]["row"],
                row_end=rows[-1]["row"],
                rows=rows,
                part_label="full_sheet",
            )
            if len(full_text.encode("utf-8")) <= max_file_bytes:
                output_path = workbook_dir / build_sheet_filename(sheet_index, sheet.title, ascii_names)
                output_path.write_text(full_text, encoding="utf-8")
                entries.append(make_manifest_entry(file_path, department, report_date, sheet.title, sheet_state, rows, output_path, None))
                continue

            sheet_dir = workbook_dir / build_sheet_stem(sheet_index, sheet.title, ascii_names)
            sheet_dir.mkdir(parents=True, exist_ok=True)
            parts = split_rows_for_size(
                rows=rows,
                source_file=file_path,
                department=department,
                report_date=report_date,
                sheet_name=sheet.title,
                sheet_state=sheet_state,
                max_file_bytes=max_file_bytes,
                overlap_rows=split_overlap_rows,
            )
            for part_index, part_rows in enumerate(parts, start=1):
                row_start = part_rows[0]["row"]
                row_end = part_rows[-1]["row"]
                output_path = sheet_dir / f"part_{part_index:03d}_rows_{row_start:04d}_{row_end:04d}.txt"
                output_path.write_text(
                    render_text(
                        source_file=file_path,
                        department=department,
                        report_date=report_date,
                        sheet_name=sheet.title,
                        sheet_state=sheet_state,
                        row_start=row_start,
                        row_end=row_end,
                        rows=part_rows,
                        part_label=f"part_{part_index:03d}",
                    ),
                    encoding="utf-8",
                )
                entries.append(make_manifest_entry(file_path, department, report_date, sheet.title, sheet_state, part_rows, output_path, part_index))
    finally:
        workbook.close()
    return entries


def make_manifest_entry(
    source_file: Path,
    department: str,
    report_date: str,
    sheet_name: str,
    sheet_state: str,
    rows: list[dict[str, Any]],
    output_path: Path,
    part_index: int | None,
) -> dict[str, Any]:
    return {
        "source_file": str(source_file),
        "source_name": source_file.name,
        "department": department,
        "report_date": report_date,
        "sheet_name": sheet_name,
        "sheet_state": sheet_state,
        "row_start": rows[0]["row"],
        "row_end": rows[-1]["row"],
        "part_index": part_index,
        "txt_file": str(output_path),
    }


def split_rows_for_size(
    rows: list[dict[str, Any]],
    source_file: Path,
    department: str,
    report_date: str,
    sheet_name: str,
    sheet_state: str,
    max_file_bytes: int,
    overlap_rows: int,
) -> list[list[dict[str, Any]]]:
    parts: list[list[dict[str, Any]]] = []
    current: list[dict[str, Any]] = []
    overlap_rows = max(0, overlap_rows)

    for row in rows:
        candidate = current + [row]
        candidate_text = render_text(
            source_file=source_file,
            department=department,
            report_date=report_date,
            sheet_name=sheet_name,
            sheet_state=sheet_state,
            row_start=candidate[0]["row"],
            row_end=candidate[-1]["row"],
            rows=candidate,
            part_label="size_probe",
        )
        if current and len(candidate_text.encode("utf-8")) > max_file_bytes:
            parts.append(current)
            current = current[-overlap_rows:] if overlap_rows else []
        current.append(row)

    if current:
        parts.append(current)
    return parts


def build_workbook_dir(output_dir: Path, file_index: int, file_path: Path, department: str, report_date: str, ascii_names: bool) -> Path:
    department_dir = f"department_{file_index:04d}" if ascii_names else safe_name(department or "unknown_department")
    document_dir = f"doc_{file_index:04d}" if ascii_names else safe_name(file_path.stem)
    if report_date:
        document_dir = f"{report_date}__{document_dir}"
    return output_dir / department_dir / document_dir


def build_sheet_stem(sheet_index: int, sheet_name: str, ascii_names: bool) -> str:
    if ascii_names:
        return f"sheet_{sheet_index:02d}"
    return f"{sheet_index:02d}_{safe_name(sheet_name)}"


def build_sheet_filename(sheet_index: int, sheet_name: str, ascii_names: bool) -> str:
    return build_sheet_stem(sheet_index, sheet_name, ascii_names) + ".txt"


def extract_rows(sheet: Any) -> list[dict[str, Any]]:
    merged_values = build_merged_value_map(sheet)
    rows: list[dict[str, Any]] = []
    for row in sheet.iter_rows():
        cells: list[dict[str, str]] = []
        for cell in row:
            value = merged_values.get(cell.coordinate)
            if value is None and not isinstance(cell, MergedCell):
                value = cell.value
            text = normalize_cell_value(value)
            if text:
                cells.append({"cell": cell.coordinate, "column": get_column_letter(cell.column), "value": text})
        if cells:
            rows.append({"row": row[0].row, "cells": cells})
    return rows


def build_merged_value_map(sheet: Any) -> dict[str, Any]:
    values: dict[str, Any] = {}
    for merged_range in sheet.merged_cells.ranges:
        top_left = sheet.cell(merged_range.min_row, merged_range.min_col)
        value = top_left.value
        if value is None:
            continue
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                values[f"{get_column_letter(col)}{row}"] = value
    return values


def render_text(
    source_file: Path,
    department: str,
    report_date: str,
    sheet_name: str,
    sheet_state: str,
    row_start: int,
    row_end: int,
    rows: list[dict[str, Any]],
    part_label: str,
) -> str:
    lines = [
        f"SOURCE_FILE: {source_file}",
        f"SOURCE_NAME: {source_file.name}",
        f"DEPARTMENT: {department or 'unknown'}",
        f"REPORT_DATE: {report_date or 'unknown'}",
        f"SHEET_NAME: {sheet_name}",
        f"SHEET_STATE: {sheet_state}",
        f"ROW_RANGE: {row_start}-{row_end}",
        f"PART: {part_label}",
        "",
        f"TITLE: {sheet_name}",
        "",
    ]
    for row in rows:
        row_text = " / ".join(cell["value"] for cell in row["cells"])
        lines.extend(
            [
                f"ROW {row['row']}",
                f"CONTEXT: source_name={source_file.name}; department={department or 'unknown'}; report_date={report_date or 'unknown'}; sheet_name={sheet_name}; row={row['row']}",
                f"TEXT: {row_text}",
            ]
        )
        for cell in row["cells"]:
            lines.append(f"CELL {cell['cell']}: {cell['value']}")
        lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def normalize_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    text = str(value)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def infer_department(file_path: Path) -> str:
    for parent in file_path.parents:
        name = parent.name
        if name and name not in {"Input Data_전처리", "Input Data", "#1"}:
            if name.endswith(("본부", "부문", "팀", "실", "센터")):
                return name
    return file_path.parent.name


def infer_report_date(filename: str) -> str:
    match = re.search(r"\((\d{6})\)", filename)
    if not match:
        return ""
    value = match.group(1)
    try:
        return date(2000 + int(value[:2]), int(value[2:4]), int(value[4:6])).isoformat()
    except ValueError:
        return ""


def safe_name(value: str) -> str:
    value = re.sub(r"[\\/:*?\"<>|]+", "_", str(value))
    value = re.sub(r"\s+", "_", value).strip("._ ")
    return value[:120] or "untitled"


def count_term(output_dir: Path, term: str) -> int:
    return sum(path.read_text(encoding="utf-8").count(term) for path in output_dir.rglob("*.txt"))


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")
    raise SystemExit(main())
