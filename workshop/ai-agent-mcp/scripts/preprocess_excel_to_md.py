from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Convert Excel workbooks into deterministic Markdown files for KB ingestion."
    )
    parser.add_argument("--input", required=True, help="Input xlsx file or directory.")
    parser.add_argument("--output", required=True, help="Output directory for generated Markdown.")
    parser.add_argument("--include-hidden", action="store_true", help="Include hidden and veryHidden sheets.")
    parser.add_argument("--max-block-rows", type=int, default=80, help="Maximum non-empty rows per Markdown block.")
    parser.add_argument("--blank-break", type=int, default=2, help="Split a block after this many blank rows.")
    parser.add_argument(
        "--ascii-names",
        action="store_true",
        help="Use ASCII-only output paths while preserving original names inside Markdown metadata.",
    )
    parser.add_argument(
        "--group-by-source",
        action="store_true",
        help="Write Markdown under department/document folders for easier S3 source tracing.",
    )
    parser.add_argument(
        "--verify-term",
        action="append",
        default=[],
        help="Term to count in generated Markdown. Can be specified multiple times.",
    )
    args = parser.parse_args()

    input_path = Path(args.input).resolve()
    output_dir = Path(args.output).resolve()
    files = collect_xlsx_files(input_path)
    if not files:
        raise SystemExit(f"No .xlsx files found: {input_path}")

    output_dir.mkdir(parents=True, exist_ok=True)
    manifest: list[dict[str, Any]] = []
    for file_index, file_path in enumerate(files, start=1):
        manifest.extend(
            convert_workbook(
                file_index=file_index,
                file_path=file_path,
                output_dir=output_dir,
                include_hidden=args.include_hidden,
                max_block_rows=args.max_block_rows,
                blank_break=args.blank_break,
                ascii_names=args.ascii_names,
                group_by_source=args.group_by_source,
            )
        )

    manifest_path = output_dir / "manifest.json"
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"workbooks={len(files)}")
    print(f"markdown_files={len(manifest)}")
    print(f"manifest={manifest_path}")
    for term in args.verify_term:
        count = count_term(output_dir, term)
        print(f"verify_term[{term}]={count}")
    return 0


def collect_xlsx_files(input_path: Path) -> list[Path]:
    if input_path.is_file() and input_path.suffix.lower() == ".xlsx":
        return [input_path]
    if input_path.is_dir():
        return sorted(
            path
            for path in input_path.rglob("*.xlsx")
            if not path.name.startswith("~$")
        )
    return []


def convert_workbook(
    file_index: int,
    file_path: Path,
    output_dir: Path,
    include_hidden: bool,
    max_block_rows: int,
    blank_break: int,
    ascii_names: bool,
    group_by_source: bool,
) -> list[dict[str, Any]]:
    workbook = load_workbook(file_path, data_only=True, read_only=False)
    department = infer_department(file_path)
    report_date = infer_report_date(file_path.name)
    workbook_dir = build_workbook_dir(
        output_dir=output_dir,
        file_index=file_index,
        file_path=file_path,
        department=department,
        report_date=report_date,
        ascii_names=ascii_names,
        group_by_source=group_by_source,
    )
    workbook_dir.mkdir(parents=True, exist_ok=True)

    entries: list[dict[str, Any]] = []
    try:
        for sheet_index, sheet in enumerate(workbook.worksheets, start=1):
            sheet_state = getattr(sheet, "sheet_state", "visible")
            if sheet_state != "visible" and not include_hidden:
                continue

            rows = extract_rows(sheet)
            blocks = split_rows(rows, max_block_rows=max_block_rows, blank_break=blank_break)
            for block_index, block in enumerate(blocks, start=1):
                row_start = block[0]["row"]
                row_end = block[-1]["row"]
                if ascii_names:
                    filename = (
                        f"sheet_{sheet_index:02d}"
                        f"_rows_{row_start:04d}_{row_end:04d}"
                        f"_part_{block_index:02d}.md"
                    )
                else:
                    filename = (
                        f"{sheet_index:02d}_{safe_name(sheet.title)}"
                        f"_rows_{row_start:04d}_{row_end:04d}"
                        f"_part_{block_index:02d}.md"
                    )
                output_path = workbook_dir / filename
                markdown = render_markdown(
                    source_file=file_path,
                    department=department,
                    report_date=report_date,
                    sheet_name=sheet.title,
                    sheet_state=sheet_state,
                    row_start=row_start,
                    row_end=row_end,
                    rows=block,
                )
                output_path.write_text(markdown, encoding="utf-8")
                entries.append(
                    {
                        "source_file": str(file_path),
                        "department": department,
                        "report_date": report_date,
                        "sheet_name": sheet.title,
                        "sheet_state": sheet_state,
                        "row_start": row_start,
                        "row_end": row_end,
                        "markdown_file": str(output_path),
                    }
                )
    finally:
        workbook.close()
    return entries




def build_workbook_dir(
    output_dir: Path,
    file_index: int,
    file_path: Path,
    department: str,
    report_date: str,
    ascii_names: bool,
    group_by_source: bool,
) -> Path:
    if ascii_names:
        document_dir = f"doc_{file_index:04d}"
    else:
        document_dir = safe_name(file_path.stem)
    if report_date:
        document_dir = f"{report_date}__{document_dir}"

    if not group_by_source:
        return output_dir / document_dir

    department_dir = safe_name(department or "unknown_department")
    return output_dir / department_dir / document_dir

def extract_rows(sheet: Any) -> list[dict[str, Any]]:
    merged_values = build_merged_value_map(sheet)
    rows: list[dict[str, Any]] = []
    for row in sheet.iter_rows():
        cells = []
        for cell in row:
            value = merged_values.get(cell.coordinate)
            if value is None and not isinstance(cell, MergedCell):
                value = cell.value
            text = normalize_cell_value(value)
            if text:
                cells.append(
                    {
                        "cell": cell.coordinate,
                        "column": get_column_letter(cell.column),
                        "value": text,
                    }
                )
        if cells:
            rows.append({"row": row[0].row, "cells": cells})
    return rows


def build_merged_value_map(sheet: Any) -> dict[str, Any]:
    values: dict[str, Any] = {}
    for merged_range in sheet.merged_cells.ranges:
        top_left = sheet.cell(merged_range.min_row, merged_range.min_col)
        value = top_left.value
        if value is None:
            continue
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                coordinate = f"{get_column_letter(col)}{row}"
                values[coordinate] = value
    return values


def split_rows(rows: list[dict[str, Any]], max_block_rows: int, blank_break: int) -> list[list[dict[str, Any]]]:
    if not rows:
        return []

    blocks: list[list[dict[str, Any]]] = []
    current: list[dict[str, Any]] = []
    previous_row: int | None = None
    for row in rows:
        gap = 0 if previous_row is None else row["row"] - previous_row - 1
        should_split = bool(current) and (
            len(current) >= max_block_rows
            or gap >= blank_break
        )
        if should_split:
            blocks.append(current)
            current = []
        current.append(row)
        previous_row = row["row"]
    if current:
        blocks.append(current)
    return blocks


def render_markdown(
    source_file: Path,
    department: str,
    report_date: str,
    sheet_name: str,
    sheet_state: str,
    row_start: int,
    row_end: int,
    rows: list[dict[str, Any]],
) -> str:
    lines = [
        "---",
        f"source_file: {yaml_escape(str(source_file))}",
        f"source_name: {yaml_escape(source_file.name)}",
        f"department: {yaml_escape(department)}",
        f"report_date: {yaml_escape(report_date)}",
        f"sheet_name: {yaml_escape(sheet_name)}",
        f"sheet_state: {yaml_escape(sheet_state)}",
        f"row_range: {row_start}-{row_end}",
        "---",
        "",
        f"# {sheet_name}",
        "",
        f"- Source file: {source_file.name}",
        f"- Department: {department or 'unknown'}",
        f"- Report date: {report_date or 'unknown'}",
        f"- Sheet: {sheet_name}",
        f"- Rows: {row_start}-{row_end}",
        "",
    ]
    for row in rows:
        row_text = " / ".join(cell["value"] for cell in row["cells"])
        lines.extend(
            [
                f"## Row {row['row']}",
                "",
                f"Context: source_name={source_file.name}; department={department or 'unknown'}; report_date={report_date or 'unknown'}; sheet_name={sheet_name}; row={row['row']}",
                f"Row text: {row_text}",
                "",
            ]
        )
        for cell in row["cells"]:
            lines.append(f"- {cell['cell']}: {cell['value']}")
        lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def normalize_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    text = str(value)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def infer_department(file_path: Path) -> str:
    for parent in file_path.parents:
        name = parent.name
        if name and name not in {"Input Data_전처리", "Input Data", "#1"}:
            if name.endswith(("본부", "부문", "팀", "실", "센터")):
                return name
    return file_path.parent.name


def infer_report_date(filename: str) -> str:
    match = re.search(r"\((\d{6})\)", filename)
    if not match:
        return ""
    value = match.group(1)
    year = int(value[:2])
    full_year = 2000 + year
    month = int(value[2:4])
    day = int(value[4:6])
    try:
        return date(full_year, month, day).isoformat()
    except ValueError:
        return ""


def safe_name(value: str) -> str:
    value = re.sub(r"[\\/:*?\"<>|]+", "_", str(value))
    value = re.sub(r"\s+", "_", value).strip("._ ")
    return value[:120] or "untitled"


def yaml_escape(value: str) -> str:
    return json.dumps(value, ensure_ascii=False)


def count_term(output_dir: Path, term: str) -> int:
    count = 0
    for path in output_dir.rglob("*.md"):
        count += path.read_text(encoding="utf-8").count(term)
    return count


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")
    raise SystemExit(main())



